import os
import PyPDF2
import re
from selenium.webdriver.common.by import By
from selenium import webdriver
import time
import openpyxl
import datetime
pdf_dir = r'C:\Users\chaitalibothe\Desktop\InvoiceProject\project invoice'
now = datetime.datetime.now()
auditlog = f"audit_{now.strftime('%Y-%m-%d_%H-%M-%S')}.txt"
with open(auditlog, "a") as file:
    for filename in os.listdir(pdf_dir):
        if filename.endswith('.pdf'):
            with open(os.path.join(pdf_dir, filename), 'rb') as pdf_file:
                pdf_reader = PyPDF2.PdfReader(pdf_file)
                num_pages = len(pdf_reader.pages)
                file.write(f"{now}: pdf data extracted\n")
                for page_num in range(num_pages):
                    page = pdf_reader.pages[page_num]
                    text = page.extract_text()
                    acc_name = re.search(r'Account Name:\s(\w+\s+\w+)',text).group(1)
                    file.write(f"{now}: vendor name extracted from pdf file\n")
                    invoice_number = re.search(r'Invoice No.:\s+(\w+)', text).group(1)
                    file.write(f"{now}: invoice name extracted\n")
                    invoice_vat_id = re.search(r' VAT Id.:\s+(\w+\s+\w+)', text).group(1)
                    file.write(f"{now}: vat id extracted\n")
                    invoice_cust_id = re.search(r'Customer Id.:\s+(\w+)', text).group(1)
                    file.write(f"{now}: customer id extracted\n")
                    invoice_date = re.search(r'Date:\s+(\w+\s+\w+\s+\w+)', text).group(1)
                    file.write(f"{now}: invoice date extracted\n")
                    invoice_bill_add = re.search(r'Billing Address:\s(\w+\s+\w+\s+\w+\s+\w+,\s+\w+)', text).group(1)
                    file.write(f"{now}: billing address extracted\n")
                    invoice_ship_add = re.search(r'Shipping Address:\s(\w+\s+\w+\s+\w+\s+\w+,\s+\w+)', text).group(1)
                    file.write(f"{now}: shipping address extracted\n")
                    invoice_sub_total = re.search(r'Sub total:\s+(\w+\s+\w+)', text).group(1)
                    file.write(f"{now}: subtotal rupees\n")
                    invoice_tax = re.search(r'Tax:\s+(\w+\s+\w+)', text).group(1)
                    file.write(f"{now}: tax rupees \n")
                    invoice_total = re.search(r'Total:\s+(\w+\s+\w+)', text).group(1)
                    file.write(f"{now}: total rupees\n")
                    invoice_gst_num = re.search(r'GST NO:\s+(\w+)', text).group(1)
                    file.write(f"{now}: gst number\n")
                    driver = webdriver.Chrome()
                    driver.get('https://form.jotform.com/230871953129461')
                    time.sleep(10)
                    file.write(f"{now}: website opened\n")
                    vendor_name =driver.find_element(By.ID, "input_6").send_keys(acc_name)
                    time.sleep(2)
                    file.write(f"{now}: vendor name entered \n")
                    invoice_num = driver.find_element(By.ID, "input_25").send_keys(invoice_number)
                    time.sleep(2)
                    file.write(f"{now}: invoice name entered \n")
                    vat_id = driver.find_element(By.ID, "input_50").send_keys(invoice_vat_id)
                    time.sleep(2)
                    file.write(f"{now}: vat id entered \n")
                    cust_id = driver.find_element(By.ID, "input_51").send_keys(invoice_cust_id)
                    time.sleep(2)
                    file.write(f"{now}: customer id entered \n")
                    date = driver.find_element(By.ID, "input_52").send_keys(invoice_date)
                    time.sleep(2)
                    file.write(f"{now}: date entered \n")
                    bill_add = driver.find_element(By.ID, "input_53").send_keys(invoice_bill_add)
                    time.sleep(2)
                    file.write(f"{now}: billing address entered \n")
                    ship_add = driver.find_element(By.ID, "input_54").send_keys(invoice_ship_add)
                    time.sleep(2)
                    file.write(f"{now}: shipping address entered \n")
                    sub_total = driver.find_element(By.ID, "input_55").send_keys(invoice_sub_total)
                    time.sleep(2)
                    file.write(f"{now}: subtotal entered \n")
                    tax = driver.find_element(By.ID, "input_56").send_keys(invoice_tax)
                    time.sleep(2)
                    file.write(f"{now}: tax rupees entered \n")
                    total = driver.find_element(By.ID, "input_57").send_keys(invoice_total)
                    time.sleep(2)
                    file.write(f"{now}: total rupees entered \n")
                    gst_num = driver.find_element(By.ID, "input_58").send_keys(invoice_gst_num)
                    time.sleep(5)
                    file.write(f"{now}: gst number entered \n")
                    # submit_button = driver.find_element_by_xpath('//button[@type="submit"]')
                    # submit_button.click()
                    wb = openpyxl.load_workbook(r"C:\Users\chaitalibothe\Desktop\InvoiceProject\invoice report.xlsx")
                    ws = wb.active
                    row = ws.max_row + 1
                    file.write(f"{now}: excel sheet opened \n")
                    ws.cell(row=row, column=1).value = acc_name
                    ws.cell(row=row, column=2).value = invoice_number
                    ws.cell(row=row, column=3).value = invoice_vat_id
                    ws.cell(row=row, column=4).value = invoice_cust_id
                    ws.cell(row=row, column=5).value = invoice_date
                    ws.cell(row=row, column=6).value = invoice_bill_add
                    ws.cell(row=row, column=7).value = invoice_ship_add
                    ws.cell(row=row, column=8).value = invoice_sub_total
                    ws.cell(row=row, column=9).value = invoice_tax
                    ws.cell(row=row, column=10).value = invoice_total
                    ws.cell(row=row, column=11).value = invoice_gst_num
                    file.write(f"{now}: details entered in excel sheet \n")
                    wb.save(r"C:\Users\chaitalibothe\Desktop\InvoiceProject\invoice report.xlsx.")
                    file.write(f"{now}: END \n")
                    time.sleep(10)
                    driver.quit()
                    time.sleep(5)
